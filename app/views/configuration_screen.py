import os
import csv
import json
import locale
import hashlib
from kivy.app import App
from logger import logger
from datetime import datetime
from kivy.uix.label import Label
from kivy.uix.popup import Popup
from kivy.uix.switch import Switch
from kivy.uix.button import Button
from kivy.uix.widget import Widget
from kivy.core.window import Window
from kivy.uix.textinput import TextInput
from kivy.uix.boxlayout import BoxLayout
from kivy.uix.screenmanager import Screen
from kivy.uix.togglebutton import ToggleButton
from kivy.uix.anchorlayout import AnchorLayout
from kivy.uix.floatlayout import FloatLayout
from kivy.uix.relativelayout import RelativeLayout
from kivy.uix.behaviors import ButtonBehavior
from kivy.graphics import Color, RoundedRectangle
from kivy.uix.scrollview import ScrollView
from openpyxl import Workbook


def afficher_popup_global(message):
    layout = BoxLayout(orientation='vertical', padding=20, spacing=20)

    label = Label(
        text=message,
        size_hint=(1, 1),
        text_size=(550, None),  # Limite la largeur → provoque des retours à la ligne
        halign="center",
        valign="middle"
    )

    btn_fermer = Button(text="Fermer", size_hint=(1, None), height=50)

    popup = Popup(
        title="Exportation",
        content=layout,
        size_hint=(None, None),
        size=(600, 600),
        auto_dismiss=False
    )

    # Adapter la hauteur du label au contenu
    label.bind(texture_size=lambda instance, value: setattr(instance, 'height', value[1]))

    btn_fermer.bind(on_press=popup.dismiss)

    layout.add_widget(label)
    layout.add_widget(btn_fermer)

    popup.open()

def exporter_vers_excel():
        json_file = "donnees_budget.json"
        dossier_export = "/storage/emulated/0/Download"
    
        try:
            locale.setlocale(locale.LC_TIME, 'fr_FR.UTF-8')
        except:
            locale.setlocale(locale.LC_TIME, '')
    
        maintenant = datetime.now()
        nom_du_mois = maintenant.strftime("%B")
        annee = maintenant.year
    
        nom_fichier = f"compte_{nom_du_mois}_{annee}.xlsx"
        fichier_excel = os.path.join(dossier_export, nom_fichier)
    
        if not os.path.exists(json_file):
            afficher_popup_global("Fichier JSON introuvable.")
            return
    
        try:
            with open(json_file, "r", encoding="utf-8") as f:
                donnees = json.load(f)
        except json.JSONDecodeError:
            afficher_popup_global("Erreur de lecture du fichier JSON.")
            return
    
        wb = Workbook()
        ws_revenus = wb.active
        ws_revenus.title = "Revenus"
        ws_revenus.append(["Date", "Nom", "Montant"])
    
        for revenu in donnees.get("revenu", []):
            ws_revenus.append([
                revenu.get("date", ""),
                revenu.get("nom", ""),
                revenu.get("montant", 0)
            ])
    
        ws_charges = wb.create_sheet(title="Charges Fixes")
        ws_charges.append(["Date", "Nom", "Montant"])
        for charge in donnees.get("charges_fixe", []):
            ws_charges.append([
                charge.get("date", ""),
                charge.get("nom", ""),
                charge.get("montant", 0)
            ])
    
        ws_depenses = wb.create_sheet(title="Dépenses")
        ws_depenses.append(["Date", "Nom", "Montant"])
        for dep in donnees.get("depense", []):
            ws_depenses.append([
                dep.get("date", ""),
                dep.get("nom", ""),
                dep.get("montant", 0)
            ])
    
        ws_epargne = wb.create_sheet(title="Épargnes")
        ws_epargne.append(["Date", "Nom", "Montant"])
        for epargne in donnees.get("epargne", []):
            ws_epargne.append([
                epargne.get("date", ""),
                epargne.get("nom", ""),
                epargne.get("montant", 0)
            ])
    
        wb.save(fichier_excel)
        afficher_popup_global(f"Exportation réussie vers :\n{fichier_excel}")
        
        






class ConfigurationScreen(Screen):
    def __init__(self, **kwargs):
        super().__init__(**kwargs)
        app = App.get_running_app()  # ✔️ Une seule fois
        
        scroll = ScrollView(size_hint=(1, 1))
        self.layout = BoxLayout(orientation='vertical', padding=20, spacing=20)

        titre = Label(text="Page de Configuration", font_size=48, size_hint=(1, None), height=100, bold=True,
                            color=(0, 0, 0, 1))
        self.layout.add_widget(titre)
        
        

# === SECTION ENCADRÉE POUR PIN ===
        pin_box = BoxLayout(
            orientation='vertical',
            spacing=10,
            padding=10,
            size_hint_y=None,
            height=350  # Hauteur plus grande
        )
        
        # Ajoute un fond encadré (optionnel mais visuel)
        with pin_box.canvas.before:
            Color(0.2, 0.6, 0.86, 1)  # fond gris clair
            self.rect = RoundedRectangle(radius=[10])
        pin_box.bind(pos=self.update_rect, size=self.update_rect)
        
        # Ligne du switch
        switch_line = BoxLayout(orientation='horizontal', size_hint_y=None, height=40)
        switch_label = Label(text="Activer le code PIN", size_hint_x=0.7)
        self.switch_pin = Switch(active=False)
        self.switch_pin.bind(active=self.toggle_pin)
        switch_line.add_widget(switch_label)
        switch_line.add_widget(self.switch_pin)
        
        # Champs de saisie
        self.new_pin_input = TextInput(
            hint_text="Nouveau PIN (4 chiffres)",
            password=False,  # pour voir les chiffres
            input_filter='int',
            multiline=False,
            size_hint_y=None,
            height=80
        )
        
        self.confirm_pin_input = TextInput(
            hint_text="Confirmer PIN",
            password=False,
            input_filter='int',
            multiline=False,
            size_hint_y=None,
            height=80
        )
        
        # Bouton enregistrer
        btn_enregistrer = Button(
            text="Enregistrer le PIN",
            size_hint_y=None,
            height=80
        )
        btn_enregistrer.bind(on_press=self.changer_pin)
        
        # Ajout à la box
        pin_box.add_widget(switch_line)
        pin_box.add_widget(self.new_pin_input)
        pin_box.add_widget(self.confirm_pin_input)
        pin_box.add_widget(btn_enregistrer)
        
        # Puis ajoute cette pin_box à ton layout principal
        self.layout.add_widget(pin_box)
               

                # Encadré switches (rouge clair)
        switches_container = BoxLayout(orientation='vertical', spacing=15, padding=15, size_hint_y=None)
        switches_container.bind(minimum_height=switches_container.setter('height'))

        with switches_container.canvas.before:
            Color(1, 0.6, 0.6, 1)  # Rouge clair
            self.bg_switch = RoundedRectangle(radius=[15], pos=switches_container.pos, size=switches_container.size)
        switches_container.bind(pos=lambda *x: setattr(self.bg_switch, 'pos', switches_container.pos))
        switches_container.bind(size=lambda *x: setattr(self.bg_switch, 'size', switches_container.size))

        app = App.get_running_app()

        # Switches
        self.switch_revenus = Switch(active=app.show_total_revenus)
        self.switch_charges = Switch(active=app.show_total_charges)
        self.switch_restant_a_payer = Switch(active=app.show_restant_a_payer)
        self.switch_depenses = Switch(active=app.show_total_depenses)

        self.switch_revenus.bind(active=self.on_toggle_revenus)
        self.switch_charges.bind(active=self.on_toggle_charges)
        self.switch_restant_a_payer.bind(active=self.on_toggle_restant_a_payer)
        self.switch_depenses.bind(active=self.on_toggle_depenses)

        switches_container.add_widget(self._create_line("Afficher Revenus", self.switch_revenus))
        switches_container.add_widget(self._create_line("Afficher Charges", self.switch_charges))
        switches_container.add_widget(self._create_line("Afficher Restant à Payer", self.switch_restant_a_payer))
        switches_container.add_widget(self._create_line("Afficher Dépenses", self.switch_depenses))

        self.layout.add_widget(switches_container)

        scroll.add_widget(self.layout)
        self.add_widget(scroll)
        
        # ✅ Widget flexible pour pousser le tout vers le haut
        self.layout.add_widget(Widget(size_hint_y=1))
        
        
        
        btn_export_excel = Button(text="Exporter en Excel", size_hint=(1, None), height=100)
        btn_export_excel.bind(on_press=lambda x: exporter_vers_excel())
        self.layout.add_widget(btn_export_excel)
        
        btn_logs = Button(text="Voir les logs", size_hint=(1, None), height=100, background_color=(1, 0, 0, 1))
        btn_logs.bind(on_press=self.voir_logs)
       # self.layout.add_widget(btn_logs)

        bouton_reset = Button(
            text="Réinitialiser revenus et depenses",
            size_hint=(1, None),
            height=100,
            background_color=(1, 0.4, 0.9, 1)
        )
        bouton_reset.bind(on_press=self.confirmer_reinitialisation)
        self.layout.add_widget(bouton_reset)
        
        bouton_retour = Button(text="Retour", size_hint=(1, None), height=100, background_color=(0.2, 0.6, 0.86, 1))
        bouton_retour.bind(on_press=self.retour_page_principale)
        self.layout.add_widget(bouton_retour)
        
    

    def confirmer_reinitialisation(self, instance):
        # Créer le message
        message = Label(
            text="Voulez-vous vraiment réinitialiser toutes les données ?",
            text_size=(Window.width * 0.7, None),
            halign="center",
            valign="middle"
        )
    
        # Créer les boutons Oui / Non
        btn_oui = Button(text="Oui", size_hint=(0.5, 1))
        btn_non = Button(text="Non", size_hint=(0.5, 1))
    
        # Layout pour les boutons
        boutons = BoxLayout(spacing=10, size_hint=(1, None), height=50)
        boutons.add_widget(btn_oui)
        boutons.add_widget(btn_non)
    
        # Conteneur principal du popup
        contenu = BoxLayout(orientation='vertical', spacing=20, padding=20)
        contenu.add_widget(message)
        contenu.add_widget(boutons)
    
        # Déclaration du popup (doit être AVANT d’être utilisé dans les callbacks)
        popup = Popup(
            title="Confirmation",
            content=contenu,
            size_hint=(0.8, None),
            height=500,
            auto_dismiss=False
        )
    
        # Liaison des boutons (utilise popup défini ci-dessus)
        btn_oui.bind(on_press=lambda *a: (self.reinitialiser_donnees(instance), popup.dismiss()))
        btn_non.bind(on_press=popup.dismiss)
    
        # Ouvre la popup
        popup.open()

        
        
    def _create_line(self, label_text, widget):
        box = BoxLayout(size_hint_y=None, height=50)
        label = Label(text=label_text, color=(0,0,0,1), size_hint_x=0.7)
        widget.size_hint_x = 0.3
        box.add_widget(label)
        box.add_widget(widget)
        return box
        
    

    def on_pre_enter(self):
        self.update_button_states()
        app = App.get_running_app()
        self.switch_pin.active = app.activer_pin
        
    def update_rect(self, *args):
        self.rect.pos = args[0].pos
        self.rect.size = args[0].size
    
    def toggle_pin(self, instance, value):
        app = App.get_running_app()
        app.activer_pin = value
        app.config_data["activer_pin"] = value
        app.sauvegarder_config()
        
    def changer_pin(self, instance):
        nouveau_pin = self.new_pin_input.text.strip()
        confirmation = self.confirm_pin_input.text.strip()

        if len(nouveau_pin) != 4 or not nouveau_pin.isdigit():
            self.afficher_popup("Erreur", "Le PIN doit contenir 4 chiffres.")
            return

        if nouveau_pin != confirmation:
            self.afficher_popup("Erreur", "Les deux codes PIN ne correspondent pas.")
            return

        # Sauvegarder dans config.json
        app = App.get_running_app()
        pin_hash = hashlib.sha256(nouveau_pin.encode()).hexdigest()
        app.config_data["pin_hash"] = pin_hash
        app.sauvegarder_config()
        self.afficher_popup("Succès", "Le code PIN a été modifié.")
        self.new_pin_input.text = ""
        self.confirm_pin_input.text = ""

    def update_button_states(self):
        app = App.get_running_app()

        self.update_button(self.switch_revenus, app.show_total_revenus, "Total Revenus")
        self.update_button(self.switch_charges, app.show_total_charges, "Total Charges")
        self.update_button(self.switch_depenses, app.show_total_depenses, "Total Dépenses")
        self.update_button(self.switch_restant_a_payer, app.show_restant_a_payer, "Restant à Payer")

        self.update_main_labels_visibility()

    def update_button(self, bouton, is_shown, label_text):
        bouton.text = f"{'Masquer' if is_shown else 'Afficher'} {label_text}"
        bouton.background_color = (0, 0.6, 0, 1) if is_shown else (1, 0, 0, 1)
        
    

    def update_main_labels_visibility(self):
        app = App.get_running_app()
        principal = self.manager.get_screen("principal")

        principal.label_revenus.opacity = 1 if app.show_total_revenus else 0
        principal.label_charges.opacity = 1 if app.show_total_charges else 0
        principal.label_depenses.opacity = 1 if app.show_total_depenses else 0
        principal.total_charges_restantes_label.opacity = 1 if app.show_restant_a_payer else 0
        
    

    def on_toggle_revenus(self, instance, value):
        app = App.get_running_app()
        app.show_total_revenus = value
        self.sauvegarder_config()

    def on_toggle_charges(self, instance, value):
        app = App.get_running_app()
        app.show_total_charges = value
        self.sauvegarder_config()

    def on_toggle_restant_a_payer(self, instance, value):
        app = App.get_running_app()
        app.show_restant_a_payer = value
        self.sauvegarder_config()

    def on_toggle_depenses(self, instance, value):
        app = App.get_running_app()
        app.show_total_depenses = value
        self.sauvegarder_config()
        
    def sauvegarder_config(self):
        app = App.get_running_app()
        config = {
            "show_total_revenus": app.show_total_revenus,
            "show_total_charges": app.show_total_charges,
            "show_total_depenses": app.show_total_depenses,
            "show_restant_a_payer": app.show_restant_a_payer
        }
        with open("config.json", "w") as f:
            json.dump(config, f)

    def voir_logs(self, instance):
        logger.info("Navigation vers l'écran Journalisation")
        self.manager.current = 'logs'

    def reinitialiser_donnees(self, instance):
        if os.path.exists("donnees_budget.json"):
            with open("donnees_budget.json", "r", encoding="utf-8") as f:
                try:
                    donnees = json.load(f)
                except json.JSONDecodeError:
                    donnees = {}
    
            # Assurer une structure de base
            if not isinstance(donnees, dict):
                donnees = {}
    
            # Extraire les charges fixes existantes (ou vide)
            charges_fixe = donnees.get("charges_fixe", [])
    
            # Réécrire le fichier avec soldes réinitialisées et charges fixes conservées
            nouvelles_donnees = {
                "soldes": [],  # suppression des revenus et dépenses
                "charges_fixe": charges_fixe,
            }
    
            with open("donnees_budget.json", "w", encoding="utf-8") as f:
                json.dump(nouvelles_donnees, f, indent=4, ensure_ascii=False)
    
        # Mise à jour de l'affichage
        principal = self.manager.get_screen("principal")
        principal.total = 0
        principal.soldes = []
        principal.solde_label.text = "Solde: 0.00 €"
        principal.label_revenus.text = "Total revenus: 0.00 €"
        principal.label_depenses.text = "Total depenses: 0.00 €"
        # Ne pas toucher à label_charges
    
        logger.warning("Les données de revenus et dépenses ont été réinitialisées.")
    
    def exporter_vers_excel():
        json_file = "donnees_budget.json"
        fichier_excel = "budget_export.xlsx"
    
        if not os.path.exists(json_file):
            print("Fichier JSON introuvable.")
            return
    
        try:
            with open(json_file, "r", encoding="utf-8") as f:
                donnees = json.load(f)
        except json.JSONDecodeError:
            print("Erreur de lecture du fichier JSON.")
            return
    
        wb = Workbook()
        ws_revenus = wb.active
        ws_revenus.title = "Revenus"
        ws_revenus.append(["Date", "Nom", "Montant"])
    
        for revenu in donnees.get("revenu", []):
            ws_revenus.append([
                revenu.get("date", ""),
                revenu.get("nom", ""),
                revenu.get("montant", 0)
            ])
    
        ws_charges = wb.create_sheet(title="Charges Fixes")
        ws_charges.append(["Date", "Nom", "Montant"])
        for charge in donnees.get("charges_fixe", []):
            ws_charges.append([
                charge.get("date", ""),
                charge.get("nom", ""),
                charge.get("montant", 0)
            ])
    
        ws_depenses = wb.create_sheet(title="Dépenses")
        ws_depenses.append(["Date", "Nom", "Montant"])
        for dep in donnees.get("depense", []):
            ws_depenses.append([
                dep.get("date", ""),
                dep.get("nom", ""),
                dep.get("montant", 0)
            ])
    
        wb.save(fichier_excel)
        print(f"Exportation réussie vers {fichier_excel}")
        
    def afficher_popup(self, titre, message):
        popup = Popup(title=titre, content=Label(text=message), size_hint=(0.6, 0.4))
        popup.open()
        
    

    def retour_page_principale(self, instance):
        
        self.manager.current = "principal"
        